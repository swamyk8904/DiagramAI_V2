from openpyxl import load_workbook


def read_excel_file(file_path):
    """
    Reads all sheets from an Excel file and returns the data as text.
    Skips completely empty rows.
    """

    workbook = load_workbook(file_path, data_only=True)

    content = ""

    for sheet in workbook.sheetnames:

        ws = workbook[sheet]

        content += f"\n===== Sheet: {sheet} =====\n"

        for row in ws.iter_rows(values_only=True):

            # Skip completely empty rows
            if all(cell is None for cell in row):
                continue

            row_text = " | ".join(
                str(cell).strip() if cell is not None else ""
                for cell in row
            )

            content += row_text + "\n"

    return content